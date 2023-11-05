# Will make an exercise on this. need to apply 10% discount on all the values under C column.
# Also need to make a chart once done.
# This is just a practice to familiarize myself with the Python packages available on the Python Package Index.
# I clearly need to delve much deeper into a specific Python package if I want to familiarize
# myself with its available functionalities

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell = sheet.cell(1, 1)
print(f'Cell value of A1 is {cell.value}')
cell = sheet.cell(1, 2)
print(f'Cell value of B2 is {cell.value}')
# cell = sheet.cell(X or column, Y or row)

max_row = sheet.max_row
print(f'''Max row for this spreadsheet is {max_row}
''')

print("will reiterate the contents of C2 to C4 cells below:")
for row in range(2, max_row + 1):
    cell_column_c = sheet.cell(row, 3)
    discounted_price = cell_column_c.value * .90
    cell_column_d = sheet.cell(row, 4)
    cell_column_d.value = discounted_price

ref_values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(ref_values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')
print("Saved on transactions2.xlsx")
